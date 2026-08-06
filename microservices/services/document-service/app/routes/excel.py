"""Excel export — trainers, requirements, email logs."""
import io
import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, Response
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel

from shared.database.service import get_db

router = APIRouter()
logger = logging.getLogger(__name__)


def _to_excel(rows: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> bytes:
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        if not rows:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([str(row.get(h, "")) if row.get(h) is not None else "" for h in headers])

        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        raise Exception("openpyxl not installed. Add it to requirements.")


def _clean(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten a MongoDB document for Excel export."""
    from datetime import datetime
    result = {}
    for k, v in doc.items():
        if k in ("_id", "resume", "combined_text", "raw_text"):
            continue
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        elif isinstance(v, (list, dict)):
            result[k] = str(v)
        else:
            result[k] = v
    return result


def _first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _shortlist_rows(shortlists: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for shortlist in shortlists:
        trainers = shortlist.get("top_trainers") or []
        base = {
            "shortlist_id": shortlist.get("shortlist_id", ""),
            "requirement_id": shortlist.get("requirement_id", ""),
            "client_name": shortlist.get("client_name", ""),
            "client_email": _first_value(shortlist.get("client_email"), shortlist.get("client_contact_email")),
            "technology": _first_value(shortlist.get("technology"), shortlist.get("technology_needed"), shortlist.get("domain")),
            "training_dates": shortlist.get("training_dates", ""),
            "timeline_start": shortlist.get("timeline_start", ""),
            "timeline_end": shortlist.get("timeline_end", ""),
            "top_count": shortlist.get("top_count", len(trainers)),
            "shortlist_status": shortlist.get("status", ""),
            "shortlist_created_at": shortlist.get("created_at", ""),
            "shortlist_updated_at": shortlist.get("updated_at", ""),
        }

        if not trainers:
            rows.append(_clean(base))
            continue

        for trainer in trainers:
            row = {
                **base,
                "rank": trainer.get("rank", ""),
                "trainer_id": trainer.get("trainer_id", ""),
                "trainer_name": _first_value(trainer.get("trainer_name"), trainer.get("name")),
                "trainer_email": _first_value(trainer.get("trainer_email"), trainer.get("email")),
                "trainer_phone": _first_value(trainer.get("trainer_phone"), trainer.get("phone")),
                "linkedin": trainer.get("linkedin", ""),
                "location": trainer.get("location", ""),
                "skills": trainer.get("skills", ""),
                "technology_category": trainer.get("technology_category", ""),
                "experience_years": trainer.get("experience_years", ""),
                "score": trainer.get("score", ""),
                "match_score": trainer.get("match_score", ""),
                "pipeline_status": trainer.get("pipeline_status", ""),
                "slot_status": trainer.get("slot_status", ""),
                "toc_status": trainer.get("toc_status", ""),
                "commercial_amount": _first_value(
                    trainer.get("commercial_amount"),
                    trainer.get("trainer_commercial"),
                    trainer.get("day_rate"),
                ),
                "last_mail_type": trainer.get("last_mail_type", ""),
                "last_mailed_at": trainer.get("last_mailed_at", ""),
                "last_mail_error": trainer.get("last_mail_error", ""),
                "interview_date": trainer.get("interview_date", ""),
                "interview_link": _first_value(trainer.get("interview_link"), trainer.get("meet_link")),
                "updated_at": trainer.get("updated_at", ""),
            }
            rows.append(_clean(row))
    return rows


@router.get("/trainers")
async def export_trainers(
    limit: int = 500,
    category: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    query: Dict[str, Any] = {}
    if category:
        query["technology_category"] = {"$regex": category, "$options": "i"}
    cursor = db.trainers.find(query, {"_id": 0, "resume": 0, "combined_text": 0}).limit(limit)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "Trainers")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=trainers.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


@router.get("/requirements")
async def export_requirements(
    limit: int = 500,
    status: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    query: Dict[str, Any] = {}
    if status:
        query["status"] = status
    cursor = db.requirements.find(query, {"_id": 0}).limit(limit)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "Requirements")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=requirements.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


@router.get("/shortlists")
async def export_shortlists(
    limit: int = 500,
    requirement_id: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    query: Dict[str, Any] = {}
    if requirement_id:
        query["requirement_id"] = requirement_id
    cursor = db.shortlists.find(query, {"_id": 0}).limit(limit).sort("updated_at", -1)
    rows = _shortlist_rows([d async for d in cursor])
    try:
        xlsx = _to_excel(rows, "Shortlists")
        filename = f"shortlist_{requirement_id}.xlsx" if requirement_id else "shortlists.xlsx"
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


@router.get("/email-logs")
async def export_email_logs(
    limit: int = 500,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    cursor = db.email_logs.find({}, {"_id": 0}).limit(limit).sort("created_at", -1)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "EmailLogs")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=email_logs.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


class ExcelImportRow(BaseModel):
    rows: List[Dict[str, Any]]
    collection: str


@router.post("/import")
async def import_excel_rows(
    payload: ExcelImportRow,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """Bulk-insert rows from an Excel import into a named collection."""
    allowed = {"trainers", "requirements", "customers"}
    if payload.collection not in allowed:
        from fastapi import HTTPException
        raise HTTPException(400, f"Collection must be one of: {', '.join(allowed)}")
    if not payload.rows:
        return {"inserted": 0}
    from datetime import datetime
    now = datetime.utcnow()
    docs = [{**row, "created_at": now, "updated_at": now, "source": "excel_import"} for row in payload.rows]
    result = await db[payload.collection].insert_many(docs)
    return {"inserted": len(result.inserted_ids), "collection": payload.collection}

"""Automatic Excel business register for TrainerSync.

This module keeps a local workbook updated from MongoDB so the business data is
available in Excel without a manual export step.
"""

from __future__ import annotations

import os
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.time_utils import utc_now


WORKBOOK_FILENAME = "trainersync_business_register.xlsx"


def workbook_path() -> Path:
    configured = os.getenv("BUSINESS_EXCEL_PATH", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    return Path(__file__).resolve().parents[1] / "assets" / WORKBOOK_FILENAME


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, (list, tuple, set)):
        return ", ".join(_text(item) for item in value if _text(item))
    if isinstance(value, dict):
        return ", ".join(f"{key}: {_text(val)}" for key, val in value.items() if _text(val))
    return str(value)


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _first(doc: dict, *keys: str) -> Any:
    for key in keys:
        value = doc.get(key)
        if value not in (None, "", [], {}):
            return value
    return ""


def _month_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%Y-%m")
    except Exception:
        return "Unknown"


def _status_observation(value: str) -> str:
    status = str(value or "").strip().lower()
    if status in {"selected", "trainer_selected_auto_sent", "training_confirmed"}:
        return "Selected / active requirement"
    if status in {"rejected", "declined", "trainer_rejected_auto_sent"}:
        return "Rejected / declined"
    if status in {"interview_scheduled", "confirmed_scheduled"}:
        return "Interview scheduled"
    if status in {"sent", "contacted", "interested", "pending_review"}:
        return "In progress"
    if status in {"superseded_premature", "superseded_false_slot_selection"}:
        return "Superseded correction"
    return status.replace("_", " ").title() if status else ""


def _setup_sheet(ws, headers: Iterable[str]) -> None:
    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _append_rows(ws, rows: Iterable[Iterable[Any]]) -> None:
    for row in rows:
        ws.append([_text(value) for value in row])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            width = max(width, min(max_width, len(_text(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


async def _load_data(db) -> dict:
    requirements = await db["requirements"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    trainers = await db["trainers"].find({}, {"_id": 0}).sort("created_at", -1).to_list(50000)
    shortlists = await db["shortlists"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    decisions = await db["post_interview_decisions"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    client_pos = await db["client_purchase_orders"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    purchase_orders = await db["purchase_orders"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    invoices = await db["invoices"].find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return {
        "requirements": requirements,
        "trainers": trainers,
        "shortlists": shortlists,
        "decisions": decisions,
        "client_pos": client_pos,
        "purchase_orders": purchase_orders,
        "invoices": invoices,
    }


def _make_summary_rows(data: dict) -> list[list[Any]]:
    requirements = data["requirements"]
    trainers = data["trainers"]
    decisions = data["decisions"]
    client_pos = data["client_pos"]
    invoices = data["invoices"]
    selected = sum(1 for req in requirements if _first(req, "selected_trainer_id") or str(_first(req, "selection_status")).lower() in {"selected", "training_confirmed"})
    rejected = sum(1 for item in decisions if ((item.get("decision") or {}).get("decision") == "rejected"))
    return [
        ["Last Updated", utc_now()],
        ["Total Trainers", len(trainers)],
        ["Total Requirements", len(requirements)],
        ["Selected Requirements", selected],
        ["Rejected Decisions", rejected],
        ["Client POs", len(client_pos)],
        ["Invoices", len(invoices)],
    ]


def _make_trainer_rows(trainers: list[dict]) -> list[list[Any]]:
    rows = []
    for index, trainer in enumerate(trainers, start=1):
        domain = _first(trainer, "domain", "primary_domain", "category", "technology", "skills")
        rows.append([
            index,
            _first(trainer, "trainer_id"),
            _first(trainer, "name", "trainer_name"),
            _first(trainer, "role", "designation", "title") or "Trainer",
            domain,
            _first(trainer, "experience", "years_experience", "total_experience", "years_of_experience"),
            _first(trainer, "email", "trainer_email"),
            _first(trainer, "phone", "mobile"),
            _first(trainer, "location", "current_location", "city"),
            _first(trainer, "certifications", "certification"),
            _first(trainer, "preferred_mode", "mode", "training_mode"),
            _first(trainer, "commercials", "expected_commercials", "rate", "commercial"),
            _first(trainer, "status", "pipeline_status"),
            _first(trainer, "observation", "notes", "remarks") or _status_observation(_first(trainer, "status", "pipeline_status")),
            _first(trainer, "created_at", "updated_at"),
        ])
    return rows


def _make_requirement_rows(requirements: list[dict]) -> list[list[Any]]:
    rows = []
    for index, req in enumerate(requirements, start=1):
        status = _first(req, "selection_status", "status")
        rows.append([
            index,
            _first(req, "requirement_id"),
            _first(req, "client_name", "client_company"),
            _first(req, "client_email"),
            _first(req, "technology_needed", "domain", "job_title"),
            _first(req, "duration", "training_duration"),
            _first(req, "training_dates", "timeline_start"),
            _first(req, "daily_timing", "timing", "training_timing"),
            _first(req, "mode", "training_mode"),
            _first(req, "budget", "commercial_budget", "client_budget"),
            status,
            _first(req, "selected_trainer_name"),
            _first(req, "selected_trainer_id"),
            _first(req, "client_po_status"),
            _first(req, "invoice_status"),
            _status_observation(status),
            _first(req, "created_at", "received_at"),
        ])
    return rows


def _make_monthly_rows(data: dict) -> list[list[Any]]:
    months: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for req in data["requirements"]:
        month = _month_key(_first(req, "created_at", "received_at"))
        months[month]["requirements"] += 1
        if _first(req, "selected_trainer_id") or str(_first(req, "selection_status")).lower() in {"selected", "training_confirmed"}:
            months[month]["selected"] += 1
    for decision in data["decisions"]:
        month = _month_key(_first(decision, "created_at", "updated_at"))
        if (decision.get("decision") or {}).get("decision") == "rejected":
            months[month]["rejected"] += 1
    for po in data["client_pos"]:
        month = _month_key(_first(po, "created_at", "updated_at", "client_po_date"))
        months[month]["client_pos"] += 1
        months[month]["po_value"] += _number(_first(po, "grand_total", "total_amount"))
    for invoice in data["invoices"]:
        month = _month_key(_first(invoice, "created_at", "invoice_date"))
        months[month]["invoices"] += 1
        months[month]["invoice_value"] += _number(_first(invoice, "commercials", "grand_total", "total_amount"))

    rows = []
    for month in sorted(months.keys(), reverse=True):
        item = months[month]
        rows.append([
            month,
            int(item["requirements"]),
            int(item["selected"]),
            int(item["rejected"]),
            int(item["client_pos"]),
            int(item["invoices"]),
            round(item["po_value"], 2),
            round(item["invoice_value"], 2),
        ])
    return rows


def _make_decision_rows(decisions: list[dict]) -> list[list[Any]]:
    """Rows for post-interview decisions only (always have a Decision ID)."""
    rows = []
    for decision in decisions:
        decision_value = (decision.get("decision") or {}).get("decision") or ""
        rows.append([
            _first(decision, "decision_id"),
            _first(decision, "requirement_id"),
            _first(decision, "trainer_name"),
            _first(decision, "trainer_id"),
            _first(decision, "client_email"),
            decision_value,
            _first(decision, "status"),
            (decision.get("decision") or {}).get("reason", ""),
            _first(decision, "reply_text"),
            _first(decision, "created_at", "updated_at"),
        ])
    return rows


def _make_shortlist_pipeline_rows(shortlists: list[dict]) -> list[list[Any]]:
    """Rows for shortlist pipeline trainer statuses (selected / rejected / stopped)."""
    rows = []
    for shortlist in shortlists:
        for trainer in shortlist.get("top_trainers") or []:
            status = _first(trainer, "pipeline_status", "status")
            if status in {"selected", "rejected", "declined", "stopped_selected"}:
                rows.append([
                    _first(shortlist, "shortlist_id", "requirement_id"),
                    _first(shortlist, "requirement_id"),
                    _first(trainer, "name", "trainer_name"),
                    _first(trainer, "trainer_id"),
                    "selected" if status == "selected" else "rejected / stopped",
                    status,
                    _first(trainer, "observation", "notes", "reason"),
                    _first(trainer, "updated_at", "selected_at", "stopped_at"),
                ])
    return rows


def _make_po_rows(client_pos: list[dict], purchase_orders: list[dict]) -> list[list[Any]]:
    rows = []
    for index, po in enumerate(client_pos, start=1):
        rows.append([
            index,
            "Client PO",
            _first(po, "client_po_number", "po_number"),
            _first(po, "requirement_id"),
            _first(po, "trainer_name"),
            _first(po, "client_name", "client_company"),
            _first(po, "client_email"),
            _first(po, "client_po_date", "po_date"),
            _first(po, "status"),
            _first(po, "grand_total", "total_amount"),
            _first(po, "source"),
            _first(po, "created_at", "updated_at"),
        ])
    offset = len(rows)
    for index, po in enumerate(purchase_orders, start=offset + 1):
        req = po.get("requirement") or {}
        trainer = po.get("trainer") or {}
        rows.append([
            index,
            "Trainer PO",
            _first(po, "po_number", "po_id"),
            _first(req, "requirement_id"),
            _first(trainer, "name", "trainer_name"),
            _first(req, "client_name", "client_company"),
            _first(req, "client_email"),
            _first(po, "po_date", "created_at"),
            _first(po, "status"),
            _first(po.get("commercials") or {}, "grand_total", "total_amount"),
            _first(po, "source"),
            _first(po, "created_at", "updated_at"),
        ])
    return rows


def _make_invoice_rows(invoices: list[dict]) -> list[list[Any]]:
    rows = []
    for index, invoice in enumerate(invoices, start=1):
        req = invoice.get("requirement") or {}
        trainer = invoice.get("trainer") or {}
        commercials = invoice.get("commercials") or {}
        rows.append([
            index,
            _first(invoice, "invoice_id"),
            _first(invoice, "invoice_number"),
            _first(invoice, "status"),
            _first(req, "requirement_id") or _first(invoice, "requirement_id"),
            _first(req, "client_name", "client_company") or _first(invoice, "client_name", "client_company"),
            _first(req, "client_email") or _first(invoice, "client_email"),
            _first(trainer, "name", "trainer_name") or _first(invoice, "trainer_name"),
            _first(invoice, "client_po_number", "po_number"),
            _first(invoice, "invoice_date", "created_at"),
            _first(commercials, "subtotal", "total_amount"),
            _first(commercials, "gst_amount"),
            _first(commercials, "grand_total") or _first(invoice, "grand_total", "total_amount"),
            _first(invoice, "sent_at"),
        ])
    return rows


async def sync_business_excel(db) -> dict:
    data = await _load_data(db)
    path = workbook_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("Summary", ["Metric", "Value"], _make_summary_rows(data)),
        ("Trainers Data", [
            "No", "Trainer ID", "Trainer Name", "Role", "Domain", "Experience", "Email", "Phone",
            "Location", "Certifications", "Mode", "Commercials", "Status", "Observation", "Created/Updated",
        ], _make_trainer_rows(data["trainers"])),
        ("Requirements Data", [
            "No", "Requirement ID", "Client", "Client Email", "Domain", "Duration", "Dates",
            "Daily Timing", "Mode", "Budget", "Status", "Selected Trainer", "Selected Trainer ID",
            "Client PO Status", "Invoice Status", "Observation", "Created",
        ], _make_requirement_rows(data["requirements"])),
        ("Monthly Summary", [
            "Month", "No of Requirements", "Selected People", "Rejected People", "No of Client POs",
            "No of Invoices", "PO Value", "Invoice Value",
        ], _make_monthly_rows(data)),
        ("Post-Interview Decisions", [
            "Decision ID", "Requirement ID", "Trainer", "Trainer ID", "Client Email", "Decision",
            "Status", "Reason", "Observation / Reply", "Date",
        ], _make_decision_rows(data["decisions"])),
        ("Shortlist Pipeline", [
            "Shortlist ID", "Requirement ID", "Trainer", "Trainer ID", "Decision",
            "Pipeline Status", "Observation / Notes", "Date",
        ], _make_shortlist_pipeline_rows(data["shortlists"])),
        ("Client PO Details", [
            "No", "PO Type", "PO Number", "Requirement ID", "Trainer", "Company / Client",
            "Client Email", "PO Date", "Status", "Amount", "Source", "Created/Updated",
        ], _make_po_rows(data["client_pos"], data["purchase_orders"])),
        ("Invoice Details", [
            "No", "Invoice ID", "Invoice Number", "Status", "Requirement ID", "Company / Client",
            "Client Email", "Trainer", "PO Number", "Invoice Date", "Subtotal", "GST", "Grand Total", "Sent At",
        ], _make_invoice_rows(data["invoices"])),
    ]

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        _setup_sheet(ws, headers)
        _append_rows(ws, rows)
        _autosize(ws)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(path.parent)) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)

    return {
        "success": True,
        "path": str(path),
        "updated_at": utc_now(),
        "sheets": [item[0] for item in sheets],
        "counts": {
            "trainers": len(data["trainers"]),
            "requirements": len(data["requirements"]),
            "decisions": len(data["decisions"]),
            "shortlists": len(data["shortlists"]),
            "client_pos": len(data["client_pos"]),
            "purchase_orders": len(data["purchase_orders"]),
            "invoices": len(data["invoices"]),
        },
    }

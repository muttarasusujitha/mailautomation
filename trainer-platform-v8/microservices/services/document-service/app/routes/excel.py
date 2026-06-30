"""Excel export — trainers, requirements, email logs."""
import io
import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, Response
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel

from app.database import get_db

router = APIRouter()
logger = logging.getLogger(__name__)


def _to_excel(rows: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> bytes:
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        if not rows:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([str(row.get(h, "")) if row.get(h) is not None else "" for h in headers])

        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        raise Exception("openpyxl not installed. Add it to requirements.")


def _clean(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten a MongoDB document for Excel export."""
    from datetime import datetime
    result = {}
    for k, v in doc.items():
        if k in ("_id", "resume", "combined_text", "raw_text"):
            continue
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        elif isinstance(v, (list, dict)):
            result[k] = str(v)
        else:
            result[k] = v
    return result


@router.get("/trainers")
async def export_trainers(
    limit: int = 500,
    category: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    query: Dict[str, Any] = {}
    if category:
        query["technology_category"] = {"$regex": category, "$options": "i"}
    cursor = db.trainers.find(query, {"_id": 0, "resume": 0, "combined_text": 0}).limit(limit)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "Trainers")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=trainers.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


@router.get("/requirements")
async def export_requirements(
    limit: int = 500,
    status: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    query: Dict[str, Any] = {}
    if status:
        query["status"] = status
    cursor = db.requirements.find(query, {"_id": 0}).limit(limit)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "Requirements")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=requirements.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


@router.get("/email-logs")
async def export_email_logs(
    limit: int = 500,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    cursor = db.email_logs.find({}, {"_id": 0}).limit(limit).sort("created_at", -1)
    rows = [_clean(d) async for d in cursor]
    try:
        xlsx = _to_excel(rows, "EmailLogs")
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=email_logs.xlsx"},
        )
    except Exception as exc:
        return {"error": str(exc), "rows": rows}


class ExcelImportRow(BaseModel):
    rows: List[Dict[str, Any]]
    collection: str


@router.post("/import")
async def import_excel_rows(
    payload: ExcelImportRow,
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    """Bulk-insert rows from an Excel import into a named collection."""
    allowed = {"trainers", "requirements", "customers"}
    if payload.collection not in allowed:
        from fastapi import HTTPException
        raise HTTPException(400, f"Collection must be one of: {', '.join(allowed)}")
    if not payload.rows:
        return {"inserted": 0}
    from datetime import datetime
    now = datetime.utcnow()
    docs = [{**row, "created_at": now, "updated_at": now, "source": "excel_import"} for row in payload.rows]
    result = await db[payload.collection].insert_many(docs)
    return {"inserted": len(result.inserted_ids), "collection": payload.collection}
